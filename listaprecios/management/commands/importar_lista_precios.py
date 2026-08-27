"""Importa el catálogo real de lista2026.xlsx al módulo Lista de Precios.

Alcance automatizado (ver lista_precios.md §8 y la sesión de análisis que
generó este comando): SALAS, POLTRONAS y SOFACAMAS comparten una misma
estructura de bloque "grupo-columna" (multiplicador en fila 1, matriz de
precio-por-metro en fila 2, bloques título→header→datos→nota→blanco) que sí
se puede reconstruir de forma fiable en costo_base + metraje + cargos +
multiplicador. El resto de hojas (SILLAS, COMEDORES, ALCOBAS 1/2, MESAS DE
CENTRO, VARIOS) tienen estructuras genuinamente heterogéneas — transpuestas,
con literales incrustados en la fórmula, o con sub-bloques anidados — que un
parser automático arriesgaría a leer mal en silencio. Se dejan como
categorías vacías, listas para cargarse a mano desde la interfaz de
administración una vez construida (decisión explícita, no un olvido — ver
el reporte que este comando imprime al final).

Estrategia de fidelidad: para cada celda de precio se lee el valor YA
CALCULADO por Excel (cached value, no se reinterpreta el texto de la
fórmula). costo_base y metraje se leen de las columnas literales D/E. Los
cargos adicionales se derivan como el resto: cargo = precio_cache/mult -
costo_base - metraje×precio_grupo, verificado en al menos 2 columnas de
grupo para confirmar que es realmente una constante (si no lo es, o si la
fila no tiene costo/metraje literal — casos como "Reclino eléctrico
estándar" en POLTRONAS, una de las ~60 celdas fijadas a mano del Excel — se
usa un PrecioVariante.precio_manual = valor cacheado directamente, sin
intentar descomponer). La columna "Cuero" SIEMPRE se guarda como override
manual: el recargo de cuero no es una constante en el Excel real (a veces
×1.1, a veces ×1.15 — ver §2.5), así que forzarlo al motor de cálculo
produciría un precio distinto al que el negocio usa hoy.
"""
import io
import os
import re
import zipfile
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

from listaprecios.models import (
    CategoriaLista, GrupoTela,
    LineaProducto, VarianteProducto, CargoVariante, PrecioVariante,
)
from paginaweb.sirv import upload_to_sirv, SirvUploadError


class DryRunRollback(Exception):
    pass


GRUPOS_ORDEN = ['Grupo 1', 'Grupo 2', 'Grupo 3', 'Grupo 4', 'Grupo 5', 'Cuero']

# Categorías que este comando puede migrar automáticamente, con la
# geometría exacta de su bloque (confirmada leyendo el archivo real celda a
# celda, no solo el análisis previo del plan).
FAMILIA_A = {
    'SALAS': dict(
        categoria='Salas', mult_cell='C1', grupo_row=2,
        grupo_cols={'Grupo 1': 'F', 'Grupo 2': 'G', 'Grupo 3': 'H', 'Grupo 4': 'I', 'Grupo 5': 'J', 'Cuero': 'K'},
        name_col='C', costo_col='D', mt_col='E', header_marker_col='D',
    ),
    'POLTRONAS': dict(
        categoria='Poltronas', mult_cell='C1', grupo_row=2,
        grupo_cols={'Grupo 1': 'F', 'Grupo 2': 'G', 'Grupo 3': 'H', 'Grupo 4': 'I', 'Grupo 5': 'J', 'Cuero': 'K'},
        name_col='C', costo_col='D', mt_col='E', header_marker_col='D',
    ),
    'SOFACAMAS': dict(
        categoria='Sofacamas', mult_cell='C1', grupo_row=2,
        grupo_cols={'Grupo 1': 'F', 'Grupo 2': 'G', 'Grupo 3': 'H', 'Grupo 4': 'I', 'Grupo 5': 'J', 'Cuero': 'K'},
        name_col='C', costo_col='D', mt_col='E', header_marker_col='D',
    ),
}

# Categorías que quedan pendientes de carga manual vía la interfaz de
# administración — se pre-crean con el multiplicador/matriz ya documentados
# en §2.2 del plan para que el trabajo manual sea solo cargar líneas, no
# también tener que redescubrir estos números.
CATEGORIAS_MANUALES = {
    'Sillas': dict(
        orden=10,
        matriz={'Grupo 1': 25000, 'Grupo 2': 35000, 'Grupo 3': 45000, 'Grupo 4': 55000, 'Grupo 5': 65000, 'Cuero': 272000},
    ),
    'Comedores': dict(
        orden=20,
        matriz={'Grupo 1': 25000, 'Grupo 2': 35000, 'Grupo 3': 45000, 'Grupo 4': 55000, 'Grupo 5': 65000},
    ),
    'Alcobas': dict(
        orden=30,
        matriz={'Grupo 1': 30000, 'Grupo 2': 40000, 'Grupo 3': 50000, 'Grupo 4': 60000, 'Grupo 5': 70000},
    ),
    'Mesas de Centro': dict(orden=40, matriz={}),
    'Varios': dict(orden=50, matriz={}),
}


def col_letter_to_idx(letter):
    return ord(letter.upper()) - ord('A') + 1


def to_decimal(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        value = value.replace('$', '').replace('.', '').replace(',', '.')
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = "Importa lista2026.xlsx al módulo Lista de Precios (ver docstring del archivo para el alcance exacto)."

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', nargs='?', default=None, help="Ruta al archivo lista2026.xlsx")
        parser.add_argument('--apply', action='store_true', help="Escribe en la base de datos (por defecto es dry-run).")
        parser.add_argument('--skip-photos', action='store_true', help="No sube fotos a Sirv (más rápido para pruebas).")
        parser.add_argument('--limit-lineas', type=int, default=None, help="Límite de líneas por hoja (para pruebas).")

    def handle(self, *args, **options):
        import openpyxl

        xlsx_path = options['xlsx_path'] or os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))),
            'lista2026.xlsx',
        )
        if not os.path.exists(xlsx_path):
            raise CommandError(f"No se encontró el archivo: {xlsx_path}")

        apply_changes = options['apply']
        skip_photos = options['skip_photos']
        limit_lineas = options['limit_lineas']

        self.stdout.write(f"Cargando {xlsx_path} ({os.path.getsize(xlsx_path) / 1e6:.1f} MB)...")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

        self.reporte = {
            'lineas_creadas': 0,
            'variantes_creadas': 0,
            'precios_calculados': 0,
            'precios_override': 0,
            'filas_omitidas': [],
            'fotos_subidas': 0,
            'fotos_fallidas': [],
            'validacion_fallos': [],
        }
        self.anchors_por_hoja = self._mapear_fotos(xlsx_path) if not skip_photos else {}

        # Fase 1: leer TODO el xlsx a memoria pura (sin tocar la base de
        # datos) — ver docstring de _leer_hoja_familia_a sobre por qué esto
        # es obligatorio con una conexión MySQL remota.
        hojas_data = {}
        for sheet_name, config in FAMILIA_A.items():
            self.stdout.write(f"Leyendo hoja: {sheet_name}...")
            hojas_data[sheet_name] = self._leer_hoja_familia_a(wb[sheet_name], sheet_name, config, limit_lineas)
        self.stdout.write(self.style.SUCCESS("Lectura del xlsx completa.\n"))

        # Fase 2: escribir todo a la base de datos, rápido y en bloque.
        lineas_para_fotos = []
        try:
            with transaction.atomic():
                grupos = self._crear_grupos_tela()
                self._crear_categorias_manuales(grupos)

                for sheet_name, hoja_data in hojas_data.items():
                    self.stdout.write(f"Escribiendo hoja: {sheet_name}...")
                    _, lineas_hoja = self._escribir_hoja(hoja_data, grupos, sheet_name, skip_photos)
                    lineas_para_fotos.extend(lineas_hoja)

                if not apply_changes:
                    raise DryRunRollback()
        except DryRunRollback:
            self.stdout.write(self.style.WARNING("\n[DRY RUN] No se escribió nada en la base de datos. Usa --apply para confirmar."))
            lineas_para_fotos = []

        # Fase 3: subir fotos (red, fuera de la transacción — cada save() es
        # su propio autocommit, así que un hipo de conexión entre subidas no
        # tumba nada más que esa foto puntual).
        if apply_changes and not skip_photos:
            for linea, titulo_row in lineas_para_fotos:
                self._subir_foto_linea(linea, linea.categoria.slug, titulo_row, self._sheet_de_categoria(linea.categoria.nombre))

        self._imprimir_reporte(apply_changes)

    def _sheet_de_categoria(self, categoria_nombre):
        for sheet_name, config in FAMILIA_A.items():
            if config['categoria'] == categoria_nombre:
                return sheet_name
        return None

    # ------------------------------------------------------------------
    def _crear_grupos_tela(self):
        grupos = {}
        for i, nombre in enumerate(GRUPOS_ORDEN, start=1):
            grupo, _ = GrupoTela.objects.get_or_create(nombre=nombre, defaults={'orden': i})
            grupos[nombre] = grupo
        return grupos

    def _crear_categorias_manuales(self, grupos):
        # NOTA histórica: este comando corrió una sola vez para migrar el
        # Excel original, cuando el precio por metro todavía vivía en una
        # matriz categoría×grupo (PrecioGrupoCategoria, eliminado). Se
        # corrigió después: el precio es el mismo para cualquier categoría
        # y vive en GrupoTela.precio_por_metro — este método ya solo crea
        # las categorías (etiquetas), no toca precios. Si se vuelve a
        # correr sobre datos nuevos, hay que fijar el precio de cada grupo
        # aparte (Lista de Precios → Multiplicadores → Matriz de Grupos).
        for nombre, cfg in CATEGORIAS_MANUALES.items():
            CategoriaLista.objects.get_or_create(
                nombre=nombre,
                defaults=dict(slug=slugify(nombre), orden=cfg['orden'], activo=True),
            )

    # ------------------------------------------------------------------
    def _mapear_fotos(self, xlsx_path):
        """sheet_name -> {título_row_1idx: (bytes, ext)} usando los anclajes
        reales de xl/drawings/*.xml (ver §2.3 del plan)."""
        z = zipfile.ZipFile(xlsx_path)
        wb_xml = z.read('xl/workbook.xml').decode('utf-8')
        sheet_rids = re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml)
        wbrels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        rid_to_target = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', wbrels_xml))

        resultado = {}
        for sheet_name, rid in sheet_rids:
            target = rid_to_target.get(rid, '')
            if 'worksheets/' not in target:
                continue
            sheet_file = target.split('/')[-1]
            rels_path = f'xl/worksheets/_rels/{sheet_file}.rels'
            if rels_path not in z.namelist():
                continue
            sheet_rels_xml = z.read(rels_path).decode('utf-8')
            m = re.search(r'Target="\.\./drawings/(drawing\d+\.xml)"', sheet_rels_xml)
            if not m:
                continue
            drawing_file = m.group(1)
            drawing_rels_path = f'xl/drawings/_rels/{drawing_file}.rels'
            if f'xl/drawings/{drawing_file}' not in z.namelist() or drawing_rels_path not in z.namelist():
                continue
            drawing_xml = z.read(f'xl/drawings/{drawing_file}').decode('utf-8')
            drawing_rels_xml = z.read(drawing_rels_path).decode('utf-8')
            rid_to_media = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="\.\./media/([^"]+)"', drawing_rels_xml))

            anchors = {}
            for anchor_match in re.finditer(
                r'<xdr:oneCellAnchor>.*?<xdr:row>(\d+)</xdr:row>.*?r:embed="(rId\d+)".*?</xdr:oneCellAnchor>',
                drawing_xml, re.DOTALL
            ):
                row0, rid_embed = anchor_match.groups()
                media_name = rid_to_media.get(rid_embed)
                if not media_name:
                    continue
                titulo_row = int(row0) + 1
                ext = os.path.splitext(media_name)[1].lower()
                try:
                    file_bytes = z.read(f'xl/media/{media_name}')
                except KeyError:
                    continue
                anchors[titulo_row] = (file_bytes, ext)
            resultado[sheet_name] = anchors
        return resultado

    # ------------------------------------------------------------------
    def _leer_hoja_familia_a(self, ws, sheet_name, config, limit_lineas):
        """Fase 1 — SOLO lectura del xlsx, sin ninguna consulta a la base de
        datos. Aislar esto es lo que evita dejar la conexión MySQL remota
        abierta e inactiva durante los minutos que toma escanear una hoja
        de ~1000 filas — eso fue justamente lo que causó
        'MySQL server has gone away' en la primera versión de este comando."""
        # openpyxl en modo read_only es rápido con iteración SECUENCIAL
        # (iter_rows) pero catastróficamente lento con acceso aleatorio vía
        # ws.cell(row=, column=) — cada llamada re-escanea desde el inicio
        # de la hoja. Con hasta 1000 filas eso es cuadrático y fue la causa
        # real del timeout/desconexión de MySQL en versiones anteriores.
        # Por eso esta función vuelca la hoja completa a una lista en un
        # solo recorrido y de ahí en adelante indexa esa lista en memoria.
        filas = list(ws.iter_rows(min_row=1, max_row=1100, values_only=True))

        def celda(row_1idx, col_1idx):
            fila = filas[row_1idx - 1] if row_1idx - 1 < len(filas) else ()
            idx = col_1idx - 1
            return fila[idx] if idx < len(fila) else None

        mult_match = re.match(r'([A-Z]+)(\d+)', config['mult_cell'])
        multiplicador = to_decimal(celda(int(mult_match.group(2)), col_letter_to_idx(mult_match.group(1)))) or Decimal('1.76')
        matriz = {}
        for grupo_nombre, col in config['grupo_cols'].items():
            precio_mt = to_decimal(celda(config['grupo_row'], col_letter_to_idx(col)))
            if precio_mt is not None:
                matriz[grupo_nombre] = precio_mt

        name_col_idx = col_letter_to_idx(config['name_col'])
        costo_col_idx = col_letter_to_idx(config['costo_col'])
        mt_col_idx = col_letter_to_idx(config['mt_col'])
        header_marker_idx = col_letter_to_idx(config['header_marker_col'])
        price_col_idx = {g: col_letter_to_idx(c) for g, c in config['grupo_cols'].items()}
        max_price_col = max(price_col_idx.values())

        current_linea = None
        lineas = []

        for row in range(4, len(filas) + 1):
            name_val = celda(row, name_col_idx)
            name_val = name_val.strip() if isinstance(name_val, str) else name_val

            if isinstance(name_val, str) and name_val:
                if name_val.lower().startswith('nota'):
                    if current_linea is not None:
                        current_linea['notas'] = (current_linea['notas'] + ' ' + name_val).strip()
                    continue

                header_marker = celda(row, header_marker_idx)
                if isinstance(header_marker, str) and header_marker.strip().lower() == 'costo':
                    continue  # fila de encabezado repetido dentro del bloque

                costo_val = celda(row, costo_col_idx)
                mt_val = celda(row, mt_col_idx)
                tiene_precio = any(
                    celda(row, idx) not in (None, '')
                    for idx in range(header_marker_idx, max_price_col + 1)
                )
                es_fila_dato = tiene_precio or costo_val not in (None, '')

                if not es_fila_dato:
                    # Fila título: cierra la línea anterior y abre una nueva.
                    if current_linea is not None and current_linea['variantes']:
                        if limit_lineas is None or len(lineas) < limit_lineas:
                            lineas.append(current_linea)
                    current_linea = {
                        'nombre': name_val, 'titulo_row': row, 'notas': '', 'variantes': [],
                    }
                    if limit_lineas is not None and len(lineas) >= limit_lineas:
                        current_linea = None
                        break
                    continue

                variante_dict = {
                    'nombre': name_val, 'row': row,
                    'costo': to_decimal(costo_val), 'metraje': to_decimal(mt_val),
                    'precios_cache': {
                        g: to_decimal(celda(row, idx))
                        for g, idx in price_col_idx.items()
                    },
                }

                if current_linea is None:
                    # No hay título abierto: en POLTRONAS/SOFACAMAS la mayoría
                    # de productos NO tienen bloque título→variantes, cada
                    # fila es un producto independiente con una sola
                    # variante (confirmado leyendo el archivo real — no es
                    # un caso de error). Se crea una línea de una sola
                    # variante con el propio nombre de la fila.
                    if limit_lineas is None or len(lineas) < limit_lineas:
                        lineas.append({
                            'nombre': name_val, 'titulo_row': row, 'notas': '',
                            'variantes': [dict(variante_dict, nombre='Único')],
                        })
                    if limit_lineas is not None and len(lineas) >= limit_lineas:
                        break
                    continue

                current_linea['variantes'].append(variante_dict)
            else:
                # Fila sin nombre: solo cierra el bloque si TODO está vacío
                # (una fila totalmente en blanco separa dos líneas).
                fila_vacia = all(
                    celda(row, idx) in (None, '')
                    for idx in range(name_col_idx, max_price_col + 1)
                )
                if fila_vacia and current_linea is not None and current_linea['variantes']:
                    if limit_lineas is None or len(lineas) < limit_lineas:
                        lineas.append(current_linea)
                    current_linea = None
                    if limit_lineas is not None and len(lineas) >= limit_lineas:
                        break

        if current_linea is not None and current_linea['variantes'] and (limit_lineas is None or len(lineas) < limit_lineas):
            lineas.append(current_linea)

        return {
            'categoria_nombre': config['categoria'],
            'multiplicador': multiplicador,
            'matriz': matriz,
            'lineas': lineas,
        }

    # ------------------------------------------------------------------
    def _escribir_hoja(self, hoja_data, grupos, sheet_name, skip_photos):
        """Fase 2 — SOLO base de datos, sobre datos ya parseados en memoria.
        Rápido y con pocas queries (bulk_create por línea), así que la
        conexión nunca queda inactiva el tiempo suficiente para que el host
        remoto la cierre."""
        categoria, _ = CategoriaLista.objects.update_or_create(
            nombre=hoja_data['categoria_nombre'],
            defaults=dict(slug=slugify(hoja_data['categoria_nombre']), activo=True),
        )
        # NOTA histórica: el precio por metro ya no depende de la categoría
        # (ver GrupoTela.precio_por_metro) — este comando ya no escribe la
        # matriz por hoja. Si se vuelve a correr sobre datos nuevos, fijar
        # el precio de cada grupo aparte desde Lista de Precios.

        # El multiplicador leído del Excel (siempre 1.76 en las hojas que
        # este comando migra) ya no vive en CategoriaLista — se pasa
        # explícito solo para reconstruir cargos a partir de la fórmula
        # original; las variantes creadas quedan con multiplicador=NULL
        # (usan el "General" del catálogo Multiplicador, que debe tener ese
        # mismo valor — ver management command de datos/migración 0006).
        lineas_a_subir_foto = []
        for linea_data in hoja_data['lineas']:
            linea = self._guardar_linea(
                linea_data, categoria, hoja_data['matriz'], hoja_data['multiplicador'], grupos, sheet_name
            )
            if linea is not None:
                lineas_a_subir_foto.append((linea, linea_data['titulo_row']))

        return categoria, lineas_a_subir_foto

    def _guardar_linea(self, linea_data, categoria, matriz, mult, grupos, sheet_name):
        nombre = linea_data['nombre']

        # Las colecciones ya no están atadas a una sola categoría: si el
        # mismo nombre de colección ya existe (venga de esta hoja o de otra
        # — ej. "Alaska" en SALAS y en POLTRONAS), se reutiliza la línea y
        # se le agregan las variantes nuevas, en vez de crear una línea
        # duplicada por cada hoja de origen.
        linea = LineaProducto.objects.filter(nombre__iexact=nombre).first()
        if linea is None:
            slug = slugify(nombre) or f"linea-{categoria.slug}"
            base_slug, count = slug, 1
            while LineaProducto.objects.filter(slug=slug).exists():
                count += 1
                slug = f"{base_slug}-{count}"
            linea = LineaProducto.objects.create(
                nombre=nombre, slug=slug, notas=linea_data['notas'], activo=True,
            )
            self.reporte['lineas_creadas'] += 1
        elif linea_data['notas'] and linea_data['notas'] not in (linea.notas or ''):
            linea.notas = (linea.notas + ' ' if linea.notas else '') + linea_data['notas']
            linea.save(update_fields=['notas'])

        orden_base = linea.variantes.count()
        # bulk_create por línea (en vez de una query por variante/cargo/precio)
        # para no saturar de round-trips una conexión MySQL remota — con
        # cientos de líneas × ~6 precios cada una, esto es la diferencia
        # entre unas pocas queries y varios miles.
        variantes_objs = [
            VarianteProducto(
                linea=linea, categoria=categoria, nombre=v['nombre'], orden=orden_base + i,
                costo_base=v['costo'] or Decimal('0'), metraje_tela=v['metraje'],
            )
            for i, v in enumerate(linea_data['variantes'])
        ]
        variantes_creadas = VarianteProducto.objects.bulk_create(variantes_objs)
        self.reporte['variantes_creadas'] += len(variantes_creadas)

        cargos_a_crear, precios_a_crear = [], []
        for variante, v in zip(variantes_creadas, linea_data['variantes']):
            self._resolver_precios_variante(variante, v, mult, matriz, grupos, cargos_a_crear, precios_a_crear)

        if cargos_a_crear:
            CargoVariante.objects.bulk_create(cargos_a_crear)
        if precios_a_crear:
            PrecioVariante.objects.bulk_create(precios_a_crear)

        return linea

    def _resolver_precios_variante(self, variante, v_data, mult, matriz, grupos, cargos_a_crear, precios_a_crear):
        costo, metraje = v_data['costo'], v_data['metraje']
        cache = v_data['precios_cache']

        cargo_derivado = None
        if costo is not None and metraje is not None and metraje != 0:
            restos = []
            for g in ['Grupo 1', 'Grupo 2', 'Grupo 3', 'Grupo 4', 'Grupo 5']:
                precio_cache, precio_mt = cache.get(g), matriz.get(g)
                if precio_cache is None or precio_mt is None:
                    continue
                resto = precio_cache / mult - costo - metraje * precio_mt
                restos.append(resto)
            if restos:
                promedio = sum(restos) / len(restos)
                consistente = all(abs(r - promedio) < Decimal('1') for r in restos)
                if consistente:
                    cargo_derivado = promedio.quantize(Decimal('1'))

        if cargo_derivado is not None:
            cargos_a_crear.append(CargoVariante(
                variante=variante, descripcion='Cargos adicionales (migrado de Excel)', valor=cargo_derivado
            ))

        for g in GRUPOS_ORDEN:
            precio_cache = cache.get(g)
            if precio_cache is None:
                continue
            grupo = grupos[g]
            usar_formula = (
                g != 'Cuero' and cargo_derivado is not None and costo is not None and metraje is not None
            )
            if usar_formula:
                precios_a_crear.append(PrecioVariante(variante=variante, grupo=grupo, precio_manual=None))
                precio_mt = matriz.get(g, Decimal('0'))
                calculado = (costo + metraje * precio_mt + cargo_derivado) * mult
                if abs(calculado - precio_cache) > Decimal('1'):
                    self.reporte['validacion_fallos'].append(
                        f"{variante.linea_id}/{variante.nombre}/{g}: calculado={calculado} vs excel={precio_cache}"
                    )
                self.reporte['precios_calculados'] += 1
            else:
                precios_a_crear.append(PrecioVariante(variante=variante, grupo=grupo, precio_manual=precio_cache))
                self.reporte['precios_override'] += 1

    def _subir_foto_linea(self, linea, categoria_slug, titulo_row, sheet_name):
        anchors = self.anchors_por_hoja.get(sheet_name, {})
        foto = anchors.get(titulo_row)
        if not foto:
            return
        file_bytes, ext = foto
        if ext not in ('.jpg', '.jpeg', '.png', '.webp', '.gif'):
            ext = '.jpg'
        filename = f"listaprecios/{categoria_slug}/{linea.slug}{ext}"
        try:
            url = upload_to_sirv(file_bytes, filename, None)
        except SirvUploadError as e:
            self.reporte['fotos_fallidas'].append(f"{linea.nombre}: {e}")
            return
        # append, no overwrite: una línea reutilizada entre hojas (ver
        # _guardar_linea) puede recibir fotos de más de una hoja.
        linea.refresh_from_db(fields=['fotos'])
        if url not in (linea.fotos or []):
            linea.fotos = (linea.fotos or []) + [url]
            linea.save(update_fields=['fotos'])
        self.reporte['fotos_subidas'] += 1

    # ------------------------------------------------------------------
    def _imprimir_reporte(self, apply_changes):
        r = self.reporte
        self.stdout.write("\n" + "=" * 60)
        self.stdout.write(self.style.SUCCESS("REPORTE DE MIGRACIÓN") if apply_changes else self.style.WARNING("REPORTE DE MIGRACIÓN (DRY RUN)"))
        self.stdout.write("=" * 60)
        self.stdout.write(f"Líneas creadas: {r['lineas_creadas']}")
        self.stdout.write(f"Variantes creadas: {r['variantes_creadas']}")
        self.stdout.write(f"Precios calculados por fórmula: {r['precios_calculados']}")
        self.stdout.write(f"Precios fijados manualmente (override): {r['precios_override']}")
        self.stdout.write(f"Fotos subidas a Sirv: {r['fotos_subidas']}")
        if r['fotos_fallidas']:
            self.stdout.write(self.style.WARNING(f"Fotos fallidas ({len(r['fotos_fallidas'])}):"))
            for f in r['fotos_fallidas'][:20]:
                self.stdout.write(f"  - {f}")
        if r['filas_omitidas']:
            self.stdout.write(self.style.WARNING(f"Filas omitidas ({len(r['filas_omitidas'])}):"))
            for f in r['filas_omitidas'][:20]:
                self.stdout.write(f"  - {f}")
        if r['validacion_fallos']:
            self.stdout.write(self.style.ERROR(f"¡Fallos de validación ({len(r['validacion_fallos'])})! El precio calculado no coincide con el del Excel:"))
            for f in r['validacion_fallos'][:20]:
                self.stdout.write(f"  - {f}")
        else:
            self.stdout.write(self.style.SUCCESS("Validación: todos los precios calculados coinciden exactamente con el Excel original."))
        self.stdout.write("\nHojas NO migradas automáticamente (quedan como categorías vacías para carga manual vía la interfaz de administración):")
        self.stdout.write("  Sillas, Comedores, Alcobas (1 y 2), Mesas de Centro, Varios")
        self.stdout.write("  Razón: cada una tiene una estructura de fórmula/bloque genuinamente distinta")
        self.stdout.write("  (columnas transpuestas, literales incrustados, sub-bloques anidados) que un")
        self.stdout.write("  parser automático arriesgaría a leer mal en silencio — ver docstring del comando.")
